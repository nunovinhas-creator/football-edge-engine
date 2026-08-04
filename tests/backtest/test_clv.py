"""
Testes de CLV (Closing Line Value) — src/backtest/historical/clv.py,
integração em models.py/evaluator.py, métricas agregadas (metrics.py),
segmentação (segments.py) e exportação (report.py / src.evaluation.report).

Cobre: CLV positivo, CLV negativo, CLV zero, ausência de closing odd,
retrocompatibilidade (dados sem os novos campos continuam a funcionar
exatamente como antes), exportação e segmentação.
"""

import os
import shutil
import tempfile
import unittest

import pandas as pd

from src.backtest.historical.clv import (
    beat_closing_market,
    calculate_clv_absolute,
    calculate_clv_percentage,
    classify_clv,
)
from src.backtest.historical.engine import BacktestEngine
from src.backtest.historical.evaluator import evaluate_bet, evaluate_bets
from src.backtest.historical.metrics import clv_summary, summary_metrics
from src.backtest.historical.models import EvaluatedBet, HistoricalBet
from src.backtest.historical.segments import (
    all_segments,
    segment_by_bookmaker,
    segment_by_clv_classification,
)
from src.backtest.historical.staking import FlatStake


def _bet(**overrides):
    defaults = dict(
        match="A vs B",
        date="2026-01-01",
        market="HOME",
        odd=2.10,
        model_prob=0.55,
        engine_decision="BET",
        result="WIN",
    )
    defaults.update(overrides)
    return HistoricalBet(**defaults)


class TestClvFormulas(unittest.TestCase):
    """Testes unitários puros das fórmulas de src/backtest/historical/clv.py."""

    def test_positive_clv(self):
        # Odd conseguida (2.10) mais alta do que o fecho (1.90) -> CLV positivo.
        absolute = calculate_clv_absolute(2.10, 1.90)
        percentage = calculate_clv_percentage(2.10, 1.90)
        self.assertAlmostEqual(absolute, 0.20, places=4)
        self.assertAlmostEqual(percentage, 10.5263, places=3)
        self.assertEqual(classify_clv(absolute), "POSITIVE")
        self.assertTrue(beat_closing_market(absolute))

    def test_negative_clv(self):
        # Fecho (3.40) subiu acima da odd conseguida (3.00) -> CLV negativo.
        absolute = calculate_clv_absolute(3.00, 3.40)
        percentage = calculate_clv_percentage(3.00, 3.40)
        self.assertAlmostEqual(absolute, -0.40, places=4)
        self.assertAlmostEqual(percentage, -11.7647, places=3)
        self.assertEqual(classify_clv(absolute), "NEGATIVE")
        self.assertFalse(beat_closing_market(absolute))

    def test_zero_clv(self):
        absolute = calculate_clv_absolute(2.00, 2.00)
        percentage = calculate_clv_percentage(2.00, 2.00)
        self.assertEqual(absolute, 0.0)
        self.assertEqual(percentage, 0.0)
        self.assertEqual(classify_clv(absolute), "NEUTRAL")
        # Neutro conta como "não perdeu valor face ao mercado".
        self.assertTrue(beat_closing_market(absolute))

    def test_missing_closing_odd_returns_none(self):
        self.assertIsNone(calculate_clv_absolute(2.10, None))
        self.assertIsNone(calculate_clv_percentage(2.10, None))
        self.assertIsNone(classify_clv(None))
        self.assertIsNone(beat_closing_market(None))

    def test_non_positive_closing_odd_returns_none_percentage(self):
        self.assertIsNone(calculate_clv_percentage(2.10, 0.0))


class TestHistoricalBetClvFields(unittest.TestCase):
    """CLV como campos opcionais de HistoricalBet — retrocompatibilidade do parsing."""

    def test_closing_odd_and_bookmaker_default_to_none(self):
        bet = _bet()
        self.assertIsNone(bet.closing_odd)
        self.assertIsNone(bet.bookmaker)

    def test_opening_odd_property_aliases_odd(self):
        bet = _bet(odd=2.35)
        self.assertEqual(bet.opening_odd, 2.35)

    def test_from_dict_without_clv_fields_still_works(self):
        # Dicionário "antigo" (pré-CLV), exatamente como já era aceite.
        row = {
            "jogo": "Benfica vs Porto",
            "data": "2026-01-01",
            "mercado": "HOME",
            "odd": 2.0,
            "probabilidade": 0.5,
            "decisao": "BET",
            "resultado": "WIN",
        }
        bet = HistoricalBet.from_dict(row)
        self.assertIsNone(bet.closing_odd)
        self.assertIsNone(bet.bookmaker)
        self.assertEqual(bet.odd, 2.0)

    def test_from_dict_accepts_portuguese_and_english_aliases(self):
        row_pt = {
            "match": "A vs B", "date": "2026-01-01", "market": "HOME",
            "odd": 2.0, "model_prob": 0.5, "engine_decision": "BET", "result": "WIN",
            "odd_fecho": 1.85, "casa_apostas": "bet365",
        }
        bet_pt = HistoricalBet.from_dict(row_pt)
        self.assertEqual(bet_pt.closing_odd, 1.85)
        self.assertEqual(bet_pt.bookmaker, "bet365")

        row_en = {
            "match": "A vs B", "date": "2026-01-01", "market": "HOME",
            "odd": 2.0, "model_prob": 0.5, "engine_decision": "BET", "result": "WIN",
            "closing_odd": 1.85, "bookmaker": "pinnacle",
        }
        bet_en = HistoricalBet.from_dict(row_en)
        self.assertEqual(bet_en.closing_odd, 1.85)
        self.assertEqual(bet_en.bookmaker, "pinnacle")


class TestEvaluateBetClv(unittest.TestCase):
    """CLV calculado por evaluate_bet, sem afetar probability/edge/ev/kelly/stake/profit."""

    def test_positive_clv_end_to_end(self):
        bet = _bet(odd=2.10, closing_odd=1.90, bookmaker="consensus")
        result = evaluate_bet(bet)
        self.assertIsInstance(result, EvaluatedBet)
        self.assertAlmostEqual(result.clv_absolute, 0.20, places=4)
        self.assertAlmostEqual(result.clv_percentage, 10.5263, places=3)
        self.assertEqual(result.clv_classification, "POSITIVE")
        self.assertEqual(result.closing_odd, 1.90)
        self.assertEqual(result.bookmaker, "consensus")

    def test_negative_clv_end_to_end(self):
        bet = _bet(odd=3.00, closing_odd=3.40, result="LOSS")
        result = evaluate_bet(bet)
        self.assertAlmostEqual(result.clv_absolute, -0.40, places=4)
        self.assertEqual(result.clv_classification, "NEGATIVE")

    def test_zero_clv_end_to_end(self):
        bet = _bet(odd=2.00, closing_odd=2.00)
        result = evaluate_bet(bet)
        self.assertEqual(result.clv_absolute, 0.0)
        self.assertEqual(result.clv_classification, "NEUTRAL")

    def test_missing_closing_odd_end_to_end(self):
        bet = _bet()
        result = evaluate_bet(bet)
        self.assertIsNone(result.closing_odd)
        self.assertIsNone(result.clv_absolute)
        self.assertIsNone(result.clv_percentage)
        self.assertIsNone(result.clv_classification)

    def test_clv_does_not_affect_core_bet_math(self):
        """
        Retrocompatibilidade: probability/edge/ev/kelly/stake/profit devem
        ser idênticos com ou sem closing_odd/bookmaker preenchidos — CLV é
        puramente aditivo.
        """
        without_clv = evaluate_bet(_bet())
        with_clv = evaluate_bet(_bet(closing_odd=1.75, bookmaker="bet365"))

        for field in ("probability", "market_probability", "edge", "ev", "kelly", "stake", "profit", "won", "placed"):
            self.assertEqual(getattr(without_clv, field), getattr(with_clv, field))

    def test_to_dict_includes_clv_and_opening_odd(self):
        result = evaluate_bet(_bet(odd=2.10, closing_odd=1.90, bookmaker="consensus"))
        d = result.to_dict()
        self.assertEqual(d["opening_odd"], 2.10)
        self.assertEqual(d["closing_odd"], 1.90)
        self.assertEqual(d["bookmaker"], "consensus")
        self.assertAlmostEqual(d["clv_absolute"], 0.20, places=4)
        self.assertEqual(d["clv_classification"], "POSITIVE")

    def test_evaluate_bets_dataframe_has_clv_columns(self):
        bets = [
            _bet(match="G1", odd=2.10, closing_odd=1.90),
            _bet(match="G2", odd=2.00),  # sem closing_odd
        ]
        df = evaluate_bets(bets)
        for column in ("closing_odd", "clv_absolute", "clv_percentage", "clv_classification", "bookmaker"):
            self.assertIn(column, df.columns)
        self.assertTrue(pd.isna(df.loc[df["match"] == "G2", "clv_absolute"]).all())


class TestClvMetrics(unittest.TestCase):
    """Métricas agregadas de CLV (src/backtest/historical/metrics.py)."""

    def setUp(self):
        bets = [
            _bet(match="G1", odd=2.10, closing_odd=1.90, competition="Liga X",
                 bookmaker="bet365", market="HOME"),   # CLV +0.20 (POSITIVE)
            _bet(match="G2", odd=3.00, closing_odd=3.40, competition="Liga X",
                 bookmaker="bet365", market="AWAY", result="LOSS"),  # CLV -0.40 (NEGATIVE)
            _bet(match="G3", odd=2.00, closing_odd=2.00, competition="Liga Y",
                 bookmaker="pinnacle", market="HOME"),  # CLV 0.0 (NEUTRAL)
            _bet(match="G4", odd=1.80, competition="Liga Y", market="DRAW"),  # sem closing_odd
        ]
        self.df = evaluate_bets(bets)

    def test_clv_coverage_pct(self):
        summary = clv_summary(self.df)
        # 3 em 4 apostas têm closing_odd -> 75%.
        self.assertAlmostEqual(summary["clv_coverage_pct"], 75.0, places=2)

    def test_avg_and_median_clv(self):
        summary = clv_summary(self.df)
        # (0.20 - 0.40 + 0.0) / 3 = -0.0667
        self.assertAlmostEqual(summary["avg_clv_absolute"], -0.0667, places=3)
        self.assertAlmostEqual(summary["median_clv_absolute"], 0.0, places=4)

    def test_positive_negative_neutral_pct(self):
        summary = clv_summary(self.df)
        self.assertAlmostEqual(summary["clv_positive_pct"], 33.33, places=1)
        self.assertAlmostEqual(summary["clv_negative_pct"], 33.33, places=1)
        self.assertAlmostEqual(summary["clv_neutral_pct"], 33.33, places=1)

    def test_beat_market_pct_includes_neutral(self):
        summary = clv_summary(self.df)
        # POSITIVE + NEUTRAL, sobre 3 apostas com CLV calculável -> 2/3.
        self.assertAlmostEqual(summary["beat_market_pct"], 66.67, places=1)

    def test_summary_metrics_includes_clv_keys(self):
        summary = summary_metrics(self.df)
        for key in ("clv_coverage_pct", "avg_clv_absolute", "clv_positive_pct", "beat_market_pct"):
            self.assertIn(key, summary)

    def test_no_closing_odd_at_all_returns_none_not_error(self):
        bets = [_bet(match="G1", odd=2.0), _bet(match="G2", odd=2.5)]
        df = evaluate_bets(bets)
        summary = clv_summary(df)
        self.assertEqual(summary["clv_coverage_pct"], 0.0)
        self.assertIsNone(summary["avg_clv_absolute"])
        self.assertIsNone(summary["clv_positive_pct"])
        self.assertIsNone(summary["beat_market_pct"])

    def test_empty_dataframe_returns_none_not_error(self):
        summary = clv_summary(pd.DataFrame())
        self.assertEqual(summary["clv_coverage_pct"], 0.0)
        self.assertIsNone(summary["avg_clv_absolute"])

    def test_dataframe_without_clv_columns_is_backward_compatible(self):
        """
        DataFrames "antigos" (sem closing_odd/clv_*, ex. construídos
        manualmente noutros testes) não devem gerar KeyError.
        """
        df = pd.DataFrame(
            [
                {"odd": 2.0, "probability": 0.5, "edge": 0.05, "ev": 0.1,
                 "kelly": 0.1, "stake": 1.0, "won": True, "profit": 1.0},
            ]
        )
        summary = summary_metrics(df)
        self.assertEqual(summary["clv_coverage_pct"], 0.0)
        self.assertIsNone(summary["avg_clv_absolute"])
        self.assertEqual(summary["roi_pct"], 100.0)  # métricas pré-existentes continuam a funcionar


class TestClvSegmentation(unittest.TestCase):
    """Segmentação CLV positivo/negativo/neutro e por bookmaker (segments.py)."""

    def setUp(self):
        bets = [
            _bet(match="G1", odd=2.10, closing_odd=1.90, bookmaker="bet365", market="HOME"),
            _bet(match="G2", odd=2.20, closing_odd=1.95, bookmaker="bet365", market="AWAY"),
            _bet(match="G3", odd=3.00, closing_odd=3.40, bookmaker="pinnacle", market="HOME", result="LOSS"),
            _bet(match="G4", odd=2.00, closing_odd=2.00, bookmaker="pinnacle", market="DRAW"),
            _bet(match="G5", odd=1.80, market="HOME"),  # sem closing_odd -> fora do segmento CLV
        ]
        self.df = evaluate_bets(bets)

    def test_segment_by_clv_classification_counts(self):
        result = segment_by_clv_classification(self.df)
        counts = dict(zip(result["clv_classification"], result["n_bets"]))
        self.assertEqual(counts.get("POSITIVE"), 2)
        self.assertEqual(counts.get("NEGATIVE"), 1)
        self.assertEqual(counts.get("NEUTRAL"), 1)
        # A aposta sem closing_odd não pertence a nenhum grupo.
        self.assertEqual(sum(counts.values()), 4)

    def test_segment_by_bookmaker_counts(self):
        result = segment_by_bookmaker(self.df)
        counts = dict(zip(result["bookmaker"], result["n_bets"]))
        self.assertEqual(counts.get("bet365"), 2)
        self.assertEqual(counts.get("pinnacle"), 2)

    def test_segment_by_bookmaker_includes_clv_metrics(self):
        result = segment_by_bookmaker(self.df)
        row = result[result["bookmaker"] == "bet365"].iloc[0]
        self.assertAlmostEqual(row["avg_clv_absolute"], 0.225, places=3)

    def test_all_segments_includes_new_clv_segments(self):
        segments = all_segments(self.df)
        self.assertIn("by_clv_classification", segments)
        self.assertIn("by_bookmaker", segments)

    def test_no_clv_data_omits_clv_classification_segment(self):
        bets = [_bet(match="G1", odd=2.0), _bet(match="G2", odd=1.9)]
        df = evaluate_bets(bets)
        segments = all_segments(df)
        self.assertNotIn("by_clv_classification", segments)


class TestClvReportExport(unittest.TestCase):
    """Exportação (CSV/Excel/HTML) com dados de CLV — src/backtest/historical/report.py."""

    @classmethod
    def setUpClass(cls):
        bets = [
            _bet(match=f"Game {i}", odd=1.5 + i * 0.1, closing_odd=1.4 + i * 0.1,
                 competition="Liga X", bookmaker="bet365", market="HOME",
                 result="WIN" if i % 2 == 0 else "LOSS")
            for i in range(10)
        ]
        # Uma aposta sem closing_odd, para exercitar a coexistência no mesmo relatório.
        bets.append(_bet(match="Game no closing", odd=2.0, competition="Liga X", market="AWAY"))

        cls.report = BacktestEngine(staking=FlatStake(unit=1.0)).run(bets)
        cls.tmp_dir = tempfile.mkdtemp(prefix="clv_report_test_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)

    def test_all_bets_dataframe_has_clv_columns(self):
        for column in ("closing_odd", "clv_absolute", "clv_percentage", "clv_classification", "bookmaker"):
            self.assertIn(column, self.report.all_bets.columns)

    def test_global_metrics_have_clv_keys(self):
        for key in ("clv_coverage_pct", "avg_clv_absolute", "clv_positive_pct", "beat_market_pct"):
            self.assertIn(key, self.report.global_metrics)

    def test_segments_include_clv_classification_and_bookmaker(self):
        self.assertIn("by_clv_classification", self.report.segments)
        self.assertIn("by_bookmaker", self.report.segments)

    def test_to_csv_writes_clv_segment_files(self):
        output_dir = os.path.join(self.tmp_dir, "csv")
        written = self.report.to_csv(output_dir)
        self.assertIn("segment_by_clv_classification", written)
        self.assertIn("segment_by_bookmaker", written)
        for key in ("segment_by_clv_classification", "segment_by_bookmaker", "bets", "summary"):
            self.assertTrue(os.path.exists(written[key]))

        bets_csv = pd.read_csv(written["bets"])
        for column in ("closing_odd", "clv_absolute", "clv_classification"):
            self.assertIn(column, bets_csv.columns)

    def test_to_html_renders_without_error(self):
        path = os.path.join(self.tmp_dir, "report.html")
        result_path = self.report.to_html(path)
        self.assertTrue(os.path.exists(result_path))
        with open(result_path, encoding="utf-8") as fh:
            content = fh.read()
        self.assertIn("clv_classification", content)

    def test_to_excel_writes_workbook_with_clv_sheets(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl não está instalado")

        path = os.path.join(self.tmp_dir, "report.xlsx")
        result_path = self.report.to_excel(path)
        self.assertTrue(os.path.exists(result_path))
        workbook = openpyxl.load_workbook(result_path, read_only=True)
        self.assertIn("by_clv_classification", workbook.sheetnames)
        self.assertIn("by_bookmaker", workbook.sheetnames)


class TestEvaluationReportMarkdownWithClv(unittest.TestCase):
    """Exportação Markdown via src.evaluation.report.EvaluationReport."""

    @classmethod
    def setUpClass(cls):
        from src.evaluation.report import evaluate

        bets = [
            _bet(match=f"Game {i}", odd=1.5 + i * 0.1, closing_odd=1.4 + i * 0.1,
                 competition="Liga X", bookmaker="bet365", market="HOME",
                 result="WIN" if i % 2 == 0 else "LOSS")
            for i in range(10)
        ]
        cls.evaluation_report = evaluate(bets, staking=FlatStake(unit=1.0))
        cls.tmp_dir = tempfile.mkdtemp(prefix="clv_markdown_test_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)

    def test_markdown_report_includes_clv_segments(self):
        path = os.path.join(self.tmp_dir, "report.md")
        result_path = self.evaluation_report.to_markdown(path)
        with open(result_path, encoding="utf-8") as fh:
            content = fh.read()
        self.assertIn("by_clv_classification", content)
        self.assertIn("by_bookmaker", content)

    def test_full_summary_via_evaluation_metrics_includes_clv(self):
        from src.evaluation.metrics import full_summary

        summary = full_summary(self.evaluation_report.placed_bets, all_df=self.evaluation_report.all_bets)
        self.assertIn("beat_market_pct", summary)
        self.assertIn("avg_clv_percentage", summary)


if __name__ == "__main__":
    unittest.main()
